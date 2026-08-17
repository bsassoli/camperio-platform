# -*- coding: utf-8 -*-
"""Data layer: accesso Oracle (sola lettura) con fallback automatico a cache JSON (demo).
LIVE se ORA_USER/ORA_PWD/ORA_DSN sono presenti e oracledb e' installato; altrimenti DEMO."""
import os, json, re, datetime
from camperio_core.portfolio import methodology as M
from camperio_core import config as _cfg
from camperio_core.oracle.client import OracleClient

HERE = os.path.dirname(os.path.abspath(__file__))
FIXTURES = os.path.join(HERE, "fixtures")


def _default_repo():
    """Repository file fondi/ETF: sotto CAMPERIO_DATA in LIVE (spec §8), locale in DEMO."""
    c = _cfg.from_env()
    if c.live:
        return str(c.data_dir / "data" / "fondi")
    return os.path.join(HERE, "data-demo", "fondi")


REPO = os.getenv("COMITATO_REPO") or _default_repo()
SCHEMI = {"ANTASIMGEST": "Gestione patrimoniale", "ANTASIMN": "Raccolta / RTO"}

FONDI_ATTESI = [("DELTA", "DELTA UCITS"), ("DELTADEF", "Delta Defensive UCITS"),
                ("SUPERDISCOVERY", "Superdiscovery UCITS"), ("ALPHAGREEN", "Alpha Green UCITS")]
ETF_ATTESI = [("IUSA", "iShares S&P 500"), ("EUN", "iShares STOXX Europe 50"),
              ("IEEM", "iShares MSCI EM"), ("EXS3", "iShares MDAX"), ("IHYG", "iShares EUR HY Corp"),
              ("EXI1", "iShares SMI"), ("ISF", "iShares FTSE 100"), ("EXS1", "iShares DAX"),
              ("EUE", "iShares Euro Stoxx 50")]

ETF_URLS = {
 "IUSA": "https://www.ishares.com/it/investitore-privato/it/prodotti/251900/ishares-sp-500-ucits-etf-inc-fund/1506575546154.ajax?fileType=csv&fileName=IUSA_holdings&dataType=fund",
 "EUN":  "https://www.ishares.com/it/investitore-privato/it/prodotti/251929/ishares-stoxx-europe-50-ucits-etf/1506575546154.ajax?fileType=csv&fileName=EUN_holdings&dataType=fund",
 "IEEM": "https://www.ishares.com/it/investitore-privato/it/prodotti/251857/ishares-msci-emerging-markets-ucits-etf-inc-fund/1506575546154.ajax?fileType=csv&fileName=IEEM_holdings&dataType=fund",
 "EXS3": "https://www.ishares.com/it/investitore-privato/it/prodotti/251845/ishares-mdax-ucits-etf-de-fund/1506575546154.ajax?fileType=csv&fileName=EXS3_holdings&dataType=fund",
 "IHYG": "https://www.ishares.com/it/investitore-privato/it/prodotti/251843/ishares-euro-high-yield-corporate-bond-ucits-etf/1506575546154.ajax?fileType=csv&fileName=IHYG_holdings&dataType=fund",
 "EXI1": "https://www.ishares.com/it/investitori-professionali/it/prodotti/251925/ishares-smi-de-fund/1506575546154.ajax?fileType=csv&fileName=EXI1_holdings&dataType=fund",
 "ISF":  "https://www.ishares.com/it/investitori-professionali/it/prodotti/251795/ishares-ftse-100-ucits-etf-inc-fund/1506575546154.ajax?fileType=csv&fileName=ISF_holdings&dataType=fund",
 "EXS1": "https://www.ishares.com/it/investitori-professionali/it/prodotti/251464/ishares-dax-ucits-etf-de-fund/1506575546154.ajax?fileType=csv&fileName=EXS1_holdings&dataType=fund",
 "EUE":  "https://www.ishares.com/it/investitori-professionali/it/prodotti/251781/ishares-euro-stoxx-50-ucits-etf-inc-fund/1506575546154.ajax?fileType=csv&fileName=EUE_holdings&dataType=fund",
}

def _load_env_file():
    p = os.path.join(HERE, "config.env")
    if os.path.exists(p):
        for line in open(p, encoding="utf-8"):
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())
_load_env_file()

# ---------------- riconoscimento fondo+data dal contenuto Excel ----------------
def parse_fund_excel(path):
    """Legge le prime righe dell'Excel del fondo e ricava (codice_fondo, data_iso).
    Ritorna (None, None) se non riconosciuto. Robusto: non dipende dal nome file."""
    fondo = None; data_iso = None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        testo = ""
        for row in ws.iter_rows(min_row=1, max_row=8, max_col=10, values_only=True):
            for v in row:
                if v is not None:
                    testo += " " + str(v)
        wb.close()
        T = testo.upper()
        if "DISCOVERY" in T:
            fondo = "SUPERDISCOVERY"
        elif "ALPHA" in T and "GREEN" in T:
            fondo = "ALPHAGREEN"
        elif "DELTA" in T and "DEFENSIVE" in T:
            fondo = "DELTADEF"
        elif "DELTA" in T:
            fondo = "DELTA"
        m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})", testo)
        if m:
            d, mo, y = m.group(1), m.group(2), m.group(3)
            if len(y) == 2:
                y = "20" + y
            data_iso = y + "-" + mo.zfill(2) + "-" + d.zfill(2)
    except Exception as e:
        print("[parse_fund_excel]", e)
    return fondo, data_iso

# ---------------- repository file fondi/ETF ----------------
def _scan_repo():
    if not os.path.isdir(REPO):
        os.makedirs(REPO, exist_ok=True)
    return [f for f in os.listdir(REPO) if f.lower().endswith((".xlsx", ".xls", ".csv"))]

def _match(prefix, files):
    cand = sorted([f for f in files if f.upper().replace("_", "").replace("-", "").replace(" ", "").startswith(prefix)
                   and not (prefix == "DELTA" and f.upper().replace("_", "").replace("-", "").replace(" ", "").startswith("DELTADEF"))], reverse=True)
    if not cand:
        return None
    f = cand[0]
    m = re.search(r"(\d{4})[-_]?(\d{2})[-_]?(\d{2})", f)
    data = (m.group(3) + "/" + m.group(2) + "/" + m.group(1)) if m else None
    return {"file": f, "data": data}

def repo_status():
    files = _scan_repo()
    def build(lista):
        out = []
        for pref, nome in lista:
            hit = _match(pref, files)
            out.append({"nome": nome, "loaded": bool(hit),
                        "file": hit["file"] if hit else "", "data": hit["data"] if hit else None,
                        "atteso": pref + "_AAAA-MM-GG"})
        return out
    return {"fondi": build(FONDI_ATTESI), "etf": build(ETF_ATTESI)}

def save_uploads(filestorages):
    """Salva i file caricati. Per gli Excel dei fondi riconosce fondo+data dal CONTENUTO
    e li rinomina in modo canonico FONDO_AAAA-MM-GG.xlsx. Ritorna lista di dict per ogni file."""
    if not os.path.isdir(REPO):
        os.makedirs(REPO, exist_ok=True)
    risultati = []
    for fs in filestorages:
        name = os.path.basename(fs.filename or "")
        if not name or not name.lower().endswith((".xlsx", ".xls", ".csv")):
            continue
        ext = os.path.splitext(name)[1].lower()
        tmp = os.path.join(REPO, "_tmp_upload" + ext)
        try:
            fs.save(tmp)
        except Exception as e:
            risultati.append({"file": name, "fondo": None, "data": None, "originale": name, "errore": str(e)})
            continue
        fondo, data_iso = (None, None)
        if ext == ".xlsx":
            fondo, data_iso = parse_fund_excel(tmp)
        if fondo and data_iso:
            final = fondo + "_" + data_iso + ext
        elif fondo:
            final = fondo + "_" + datetime.date.today().isoformat() + ext
        else:
            final = name
        dest = os.path.join(REPO, final)
        try:
            os.replace(tmp, dest)
        except Exception:
            if os.path.exists(tmp):
                os.remove(tmp)
            risultati.append({"file": name, "fondo": fondo, "data": data_iso, "originale": name, "errore": "salvataggio"})
            continue
        risultati.append({"file": final, "fondo": fondo, "data": data_iso, "originale": name})
    return risultati

def download_etf(timeout=25):
    """Scarica i CSV holdings dei 9 ETF benchmark da iShares nel Repository_Fondi."""
    import urllib.request
    if not os.path.isdir(REPO):
        os.makedirs(REPO, exist_ok=True)
    oggi = datetime.date.today().isoformat()
    out = []
    for tk, url in ETF_URLS.items():
        dest = os.path.join(REPO, tk + "_" + oggi + ".csv")
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                data = resp.read()
            if len(data) < 200:
                out.append({"ticker": tk, "ok": False, "msg": "risposta vuota/non valida"}); continue
            with open(dest, "wb") as f:
                f.write(data)
            out.append({"ticker": tk, "ok": True, "msg": "scaricato", "file": os.path.basename(dest)})
        except Exception as e:
            out.append({"ticker": tk, "ok": False, "msg": "non raggiungibile (" + type(e).__name__ + ")"})
    return out

# ---------------- connessione Oracle ----------------
# Connessione via core: con config LIVE un Oracle irraggiungibile solleva
# OracleIndisponibileError — mai dati DEMO spacciati per veri (spec §4).
_client = OracleClient()


def mode():
    return _client.mode()


def _q(sql, params=None):
    return _client.query(sql, params)

LINEA = {"TR1": "Treasury", "TRP": "Treasury Plus", "CIN": "Income", "CPS": "Income Plus",
         "CLV": "Low Volatility", "CMR": "Camperio", "CPP": "Camperio Plus", "CME": "Equity",
         "RTO": "RTO", "PRO": "Conto proprio", "PRG": "Programmato"}

def _gruppo(schema, tipoge):
    """Gestione = ANTASIMGEST; Consulenza = ANTASIMN con TIPOGE che inizia per 'L'; RTO = altri ANTASIMN."""
    if schema == "ANTASIMGEST":
        return "Gestione"
    return "Consulenza" if (tipoge or "").upper().startswith("L") else "RTO"

def list_contratti():
    if mode() == "DEMO":
        out = json.load(open(os.path.join(FIXTURES, "contratti.json"), encoding="utf-8"))
        for c in out:
            c["gruppo"] = _gruppo(c.get("schema"), c.get("tipoge"))
        return out
    out = []
    for sch in SCHEMI:
        rows = _q("SELECT CODCLI, DESCLI, TIPOGE FROM " + sch + ".CNT WHERE DATCHIU IS NULL ORDER BY CODCLI")
        for r in rows:
            out.append({"schema": sch, "codcli": r["CODCLI"], "descli": r.get("DESCLI") or r["CODCLI"],
                        "tipoge": r.get("TIPOGE") or "", "linea": LINEA.get(r.get("TIPOGE"), r.get("TIPOGE") or ""),
                        "gruppo": _gruppo(sch, r.get("TIPOGE"))})
    return out

def _load_cache(codcli):
    p = os.path.join(FIXTURES, codcli + ".json")
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else None

def available_dates(schema, codcli):
    if mode() == "DEMO":
        pf = _load_cache(codcli)
        return pf["meta"]["date_disponibili"] if pf else []
    rows = _q("SELECT TO_CHAR(PERIODO,'YYYY-MM-DD') D FROM " + schema +
              ".SRE WHERE CODCLI=:c AND PERIODO>=ADD_MONTHS(SYSDATE,-3) ORDER BY PERIODO DESC", {"c": codcli})
    return [r["D"] for r in rows]

def _minus7(d):
    return (datetime.date.fromisoformat(d) - datetime.timedelta(days=7)).isoformat()

def get_portfolio(schema, codcli, dal=None, al=None, data=None):
    """Periodo di analisi: 'dal' (iniziale) e 'al' (finale). 'data' = retro-compat (=al)."""
    if mode() == "DEMO":
        pf = _load_cache(codcli)
        if not pf:
            raise ValueError("Nessun dato in cache per " + str(codcli) + " (DEMO: disponibile solo DEMO01).")
        return pf

    assert schema in SCHEMI, "schema non valido"
    al = al or data or _q("SELECT TO_CHAR(MAX(PERIODO),'YYYY-MM-DD') D FROM " + schema + ".SRE WHERE CODCLI=:c", {"c": codcli})[0]["D"]
    dal = dal or _minus7(al)

    def _sre_at(d):
        rows = _q("SELECT TO_CHAR(PERIODO,'YYYY-MM-DD') D, CONSFIN, TCLI, TBMK, CODIND FROM " + schema +
                  ".SRE WHERE CODCLI=:c AND PERIODO<=TO_DATE(:d,'YYYY-MM-DD') ORDER BY PERIODO DESC FETCH FIRST 1 ROWS ONLY",
                  {"c": codcli, "d": d})
        return rows[0] if rows else None

    curr = _sre_at(al); prev = _sre_at(dal) or curr
    drif = curr["D"]
    base = _q("SELECT CONSFIN, TCLI, TBMK FROM " + schema +
              ".SRE WHERE CODCLI=:c AND PERIODO<=DATE '2025-12-31' ORDER BY PERIODO DESC FETCH FIRST 1 ROWS ONLY", {"c": codcli})
    base = base[0] if base else {"CONSFIN": curr["CONSFIN"], "TCLI": curr["TCLI"], "TBMK": curr["TBMK"]}

    wp = _q("SELECT CODABI, DESTITB DES, GRUTIT, DIVISA, QUANTI, VALMER, VALOREFUT, DELTA, CODISIN, BLOOMBERG FROM " +
            schema + ".WCTDD WHERE CODCLI=:c AND VALMER<>0 ORDER BY VALMER DESC", {"c": codcli})
    positions = []
    for r in wp:
        d = dict(codabi=r["CODABI"], des=r["DES"] or "", grutit=r["GRUTIT"] or "",
                 valmer=float(r["VALMER"] or 0), valorefut=float(r["VALOREFUT"] or r["VALMER"] or 0),
                 bbg=r.get("BLOOMBERG") or "", isin=r.get("CODISIN") or "")
        mac, sub = M.macro(d["grutit"]); d["macro"] = mac; d["sub"] = sub; d["ccy"] = M.currency_of(d)
        positions.append(d)

    pod = _q("SELECT CODABI, PNET, VAL, TO_CHAR(DATSCA,'YYYY-MM-DD') SCAD FROM " + schema +
             ".POD WHERE CODCLI=:c AND DATPOS=(SELECT MAX(DATPOS) FROM " + schema + ".POD WHERE CODCLI=:c)", {"c": codcli})
    mov = _q("SELECT CODABI, TIPOPE, QUANTI, CTVTIT, TO_CHAR(DATOPE,'YYYY-MM-DD') DATA FROM " + schema +
             ".MOV WHERE CODCLI=:c AND TRUNC(DATOPE) BETWEEN TO_DATE(:p,'YYYY-MM-DD') AND TO_DATE(:d,'YYYY-MM-DD') ORDER BY DATOPE",
             {"c": codcli, "p": prev["D"], "d": drif})

    meta = {"codcli": codcli, "schema": schema, "tipoge": "", "linea": "",
            "descli": codcli, "data": drif, "data_prec": prev["D"],
            "nav": float(curr["CONSFIN"]), "date_disponibili": available_dates(schema, codcli)}
    return {
        "meta": meta, "positions": positions,
        "sre": {"base": {"nav": float(base["CONSFIN"]), "tcli": float(base["TCLI"]), "tbmk": float(base["TBMK"])},
                "prev": {"nav": float(prev["CONSFIN"]), "tcli": float(prev["TCLI"]), "tbmk": float(prev["TBMK"])},
                "curr": {"nav": float(curr["CONSFIN"]), "tcli": float(curr["TCLI"]), "tbmk": float(curr["TBMK"])},
                "codind": curr.get("CODIND") or "", "benchmark": ""},
        "pod": [{"codabi": r["CODABI"], "des": r["CODABI"], "pnet": r["PNET"], "val": r["VAL"],
                 "scad": r.get("SCAD") or "", "tipo": ""} for r in pod],
        "mov": [{"codabi": r["CODABI"], "nome": r["CODABI"], "op": r.get("TIPOPE") or "", "eff": "",
                 "q": r.get("QUANTI"), "ctv": float(r.get("CTVTIT") or 0), "data": r.get("DATA") or ""}
                for r in mov if not str(r["CODABI"]).startswith(("LIQ", "IMI", "LMS", "LMI"))],
    }

def _fmt_it(iso):
    try:
        y, m, d = iso.split("-"); return d + "/" + m + "/" + y
    except Exception:
        return iso

def resolve_period(schema, codcli, dal, al):
    """Risolve le date scelte a calendario contro le date dati disponibili.
    AL: deve esistere esattamente, altrimenti il report NON si produce.
    DAL: se non esiste, usa l'ultima disponibile <= DAL (o la piu' vecchia) e avvisa."""
    avail = available_dates(schema, codcli)  # ordine desc
    if not avail:
        return {"ok": False, "error": "Nessun dato disponibile per questo portafoglio."}
    amax = avail[0]
    al_eff = al or amax
    if al and al not in avail:
        return {"ok": False, "error": ("Alla data AL " + _fmt_it(al) + " non ci sono dati di portafoglio: "
                "il report non viene prodotto. Ultima data disponibile: " + _fmt_it(amax) + ".")}
    dal_eff = dal or _minus7(al_eff)
    warn = None
    if dal_eff not in avail:
        prev = [d for d in avail if d <= dal_eff]
        used = prev[0] if prev else avail[-1]
        warn = ("Alla data DAL " + _fmt_it(dal_eff) + " non ci sono dati: uso l'ultima disponibile "
                + _fmt_it(used) + ".")
        dal_eff = used
    return {"ok": True, "al": al_eff, "dal": dal_eff, "warn": warn}

# ---------------- Variazioni prezzi (report 3): prezzi storici VAL/TIT + cambio per data ----------------
# Codice divisa interno -> ISO (dai forward V02: 001=USD,002=GBP,003=CHF,007=DKK,008=NOK,009=SEK,071=JPY,124=SGD)
DIVI_ISO = {"000": "EUR", "": "EUR", "001": "USD", "002": "GBP", "003": "CHF",
            "007": "DKK", "008": "NOK", "009": "SEK", "071": "JPY", "124": "SGD"}

def price_changes(schema, codcli, dal, al):
    """Variazioni di prezzo dei titoli azionari diretti tra 'dal' e 'al'.
    Prezzo per azione in EUR = prezzo locale (VAL) / cambio della data.
    Cambio: forward V02 con CODABI = codice divisa (USD/GBP/CHF/DKK/NOK/SEK...);
            JPY via cross EUR-USD (009253) x USD-JPY (009254); EUR = 1.
    DEMO: legge cache/<codcli>_var.json. Ritorna {nav, rows[]} con loc/fx grezzi per data."""
    if mode() == "DEMO":
        p = os.path.join(FIXTURES, codcli + "_var.json")
        return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else {"nav": 0.0, "rows": []}
    assert schema in SCHEMI, "schema non valido"
    navr = _q("SELECT CONSFIN FROM " + schema + ".SRE WHERE CODCLI=:c AND PERIODO=TO_DATE(:d,'YYYY-MM-DD')",
              {"c": codcli, "d": al})
    nav = float(navr[0]["CONSFIN"]) if navr else 0.0
    S = schema
    sql = ("""SELECT p.codabi, p.nome, p.divi, p.qty,
      (SELECT MAX(VALMER) FROM {S}.VAL WHERE CODABI=p.codabi AND DATA=TO_DATE(:dal,'YYYY-MM-DD')) loc_dal,
      (SELECT MAX(VALMER) FROM {S}.VAL WHERE CODABI=p.codabi AND DATA=TO_DATE(:al,'YYYY-MM-DD')) loc_al,
      CASE WHEN p.divi IN ('000',' ') THEN 1
           WHEN p.divi='071' THEN (SELECT MAX(VALMER) FROM {S}.VAL WHERE CODABI='009253' AND DATA=TO_DATE(:dal,'YYYY-MM-DD'))
                                  *(SELECT MAX(VALMER) FROM {S}.VAL WHERE CODABI='009254' AND DATA=TO_DATE(:dal,'YYYY-MM-DD'))
           ELSE (SELECT MAX(v.VALMER) FROM {S}.VAL v JOIN {S}.TIT t ON t.CODABI=v.CODABI
                 WHERE t.GRUTIT='V02' AND LENGTH(t.CODABI)<=3 AND t.CODABI=p.divi AND v.DATA=TO_DATE(:dal,'YYYY-MM-DD')) END fx_dal,
      CASE WHEN p.divi IN ('000',' ') THEN 1
           WHEN p.divi='071' THEN (SELECT MAX(VALMER) FROM {S}.VAL WHERE CODABI='009253' AND DATA=TO_DATE(:al,'YYYY-MM-DD'))
                                  *(SELECT MAX(VALMER) FROM {S}.VAL WHERE CODABI='009254' AND DATA=TO_DATE(:al,'YYYY-MM-DD'))
           ELSE (SELECT MAX(v.VALMER) FROM {S}.VAL v JOIN {S}.TIT t ON t.CODABI=v.CODABI
                 WHERE t.GRUTIT='V02' AND LENGTH(t.CODABI)<=3 AND t.CODABI=p.divi AND v.DATA=TO_DATE(:al,'YYYY-MM-DD')) END fx_al
    FROM (SELECT w.CODABI codabi, MAX(t.DESTITB) nome, MAX(NVL(t.DIVI,'000')) divi, SUM(w.QUANTI) qty
          FROM {S}.WCTDD w JOIN {S}.TIT t ON t.CODABI=w.CODABI
          WHERE w.CODCLI=:c AND t.GRUTIT LIKE 'E%' AND w.QUANTI<>0 GROUP BY w.CODABI) p
    ORDER BY p.nome""").format(S=S)
    rows = _q(sql, {"c": codcli, "dal": dal, "al": al})
    def _f(v):
        return float(v) if v is not None else None
    out = []
    for r in rows:
        divi = (r["DIVI"] or "000").strip()
        out.append({"codabi": r["CODABI"], "nome": r["NOME"] or "", "divi": divi,
                    "ccy": DIVI_ISO.get(divi, divi), "qty": float(r["QTY"] or 0),
                    "loc_dal": _f(r["LOC_DAL"]), "loc_al": _f(r["LOC_AL"]),
                    "fx_dal": _f(r["FX_DAL"]), "fx_al": _f(r["FX_AL"])})
    # Quantita' alla data AL = posizione attuale (WCTDD) - ordini eseguiti dopo AL (registro ORD: solo trade, no movimenti tecnici)
    try:
        adj = _q("SELECT CODABI, SUM(CASE WHEN TIPOPE IN ('AVD','AVE','AST','ASP') THEN -QUANTI ELSE QUANTI END) net FROM "
                 + S + ".ORD WHERE CODCLI=:c AND TRUNC(DATSTI) > TO_DATE(:al,'YYYY-MM-DD') GROUP BY CODABI",
                 {"c": codcli, "al": al})
        net = {r["CODABI"]: float(r["NET"] or 0) for r in adj}
    except Exception as e:
        print("[price_changes] ORD non disponibile:", e); net = {}
    for r in out:
        r["qty_al"] = r["qty"] - net.get(r["codabi"], 0.0)
    return {"nav": nav, "rows": out, "dal": dal, "al": al}

# ---------------- Report 4 "Sintesi Comitato" (Word): performance + pesi AL/DAL + operazioni ----------------
def comitato_extra(schema, codcli, dal, al):
    """Dati aggiuntivi per la sintesi Comitato: performance (SRE TCLI/TBMK), pesi per asset class
    ad AL e DAL (effetto mercato, da WCTDD+VAL), operazioni del periodo (registro ORD, escluso FX).
    DEMO: cache/<codcli>_comitato.json."""
    if mode() == "DEMO":
        p = os.path.join(FIXTURES, codcli + "_comitato.json")
        return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else {}
    assert schema in SCHEMI, "schema non valido"
    S = schema
    def _sre(d):
        r = _q("SELECT CONSFIN, TCLI, TBMK, CODIND FROM " + S + ".SRE WHERE CODCLI=:c AND PERIODO=TO_DATE(:d,'YYYY-MM-DD')", {"c": codcli, "d": d})
        return r[0] if r else None
    sa = _sre(al); sd = _sre(dal) or sa
    bq = _q("SELECT TCLI, TBMK, CONSFIN FROM " + S + ".SRE WHERE CODCLI=:c AND PERIODO < TRUNC(TO_DATE(:al,'YYYY-MM-DD'),'YEAR') ORDER BY PERIODO DESC FETCH FIRST 1 ROWS ONLY", {"c": codcli, "al": al})
    der = _q("SELECT NVL(SUM(NVL(w.VALOREFUT,0)),0) de FROM " + S + ".WCTDD w JOIN " + S + ".TIT t ON t.CODABI=w.CODABI "
             "WHERE w.CODCLI=:c AND (t.GRUTIT LIKE 'F%' OR t.GRUTIT LIKE 'G%') AND ("
             "UPPER(t.DESTITB) LIKE '%S&P%' OR UPPER(t.DESTITB) LIKE '%MINI FUT%' OR UPPER(t.DESTITB) LIKE '%STOXX%' "
             "OR UPPER(t.DESTITB) LIKE '%DAX%' OR UPPER(t.DESTITB) LIKE '%FTSE%' OR UPPER(t.DESTITB) LIKE '%SMI %' "
             "OR UPPER(t.DESTITB) LIKE '%MSCI%' OR UPPER(t.DESTITB) LIKE '%NASDAQ%' OR UPPER(t.DESTITB) LIKE '%NIKKEI%')", {"c": codcli})
    der_eq = float(der[0]["DE"]) if der else 0.0
    dp = _q("SELECT t.DESTITB nome, t.CODISIN isin, t.GRUTIT g, NVL(w.VALOREFUT,0) vf, NVL(w.VALMER,0) vm "
            "FROM " + S + ".WCTDD w JOIN " + S + ".TIT t ON t.CODABI=w.CODABI "
            "WHERE w.CODCLI=:c AND (t.GRUTIT LIKE 'F%' OR t.GRUTIT LIKE 'G%') AND (NVL(w.VALOREFUT,0)<>0 OR NVL(w.VALMER,0)<>0) "
            "ORDER BY ABS(NVL(w.VALOREFUT,0)) DESC", {"c": codcli})
    deriv_pos = [{"nome": r["NOME"], "isin": r.get("ISIN") or "", "valorefut": float(r["VF"]), "valmer": float(r["VM"])} for r in dp]
    bb = bq[0] if bq else None
    comp = _q(("""SELECT macro, SUM(val_al) val_al, SUM(val_dal) val_dal FROM (
        SELECT CASE WHEN cab='95ZYC2' THEN 'Fondo DELTA UCITS'
                    WHEN cab='8W8C01' THEN 'Fondo Superdiscovery'
                    WHEN cab='54OPLF' THEN 'Fondo Alpha Green'
                    WHEN cab IN ('8W8C34','8W8C45','50QNF8','52PZG8') THEN 'Fondo Delta Defensive'
                    WHEN g LIKE 'A%' THEN 'Obbligazioni'
                    WHEN g='H19' THEN 'Oro fisico'
                    WHEN g IN ('H06','H20') THEN 'Altri fondi UCITS'
                    WHEN g LIKE 'E%' OR g IN ('H23','H10','H18') THEN 'Azioni ed ETF azionari'
                    WHEN g LIKE 'Z%' THEN 'Liquidita'
                    ELSE 'Derivati e altro' END macro,
               VALMER val_al, CASE WHEN pal>0 THEN VALMER*pdal/pal ELSE VALMER END val_dal
        FROM (SELECT w.CODABI cab, t.GRUTIT g, w.VALMER,
                NVL((SELECT MAX(VALMER) FROM """ + S + """.VAL WHERE CODABI=w.CODABI AND DATA=TO_DATE(:al,'YYYY-MM-DD')),0) pal,
                NVL((SELECT MAX(VALMER) FROM """ + S + """.VAL WHERE CODABI=w.CODABI AND DATA=TO_DATE(:dal,'YYYY-MM-DD')),0) pdal
              FROM """ + S + """.WCTDD w JOIN """ + S + """.TIT t ON t.CODABI=w.CODABI
              WHERE w.CODCLI=:c AND w.VALMER<>0)
        ) GROUP BY macro"""), {"c": codcli, "al": al, "dal": dal})
    trades = _q(("""SELECT t.DESTITB nome, t.GRUTIT g, o.TIPOPE,
        CASE WHEN o.TIPOPE LIKE 'AV%' OR o.TIPOPE LIKE '%VS' THEN 'Vendita'
             WHEN o.TIPOPE LIKE 'DA%' OR o.TIPOPE LIKE 'AC%' OR o.TIPOPE LIKE '%CQ%' THEN 'Acquisto'
             WHEN o.TIPOPE LIKE 'E1%' OR t.GRUTIT LIKE 'G%' THEN 'Future (roll)'
             ELSE o.TIPOPE END verso,
        o.QUANTI, o.PREZUNI, TO_CHAR(o.DATSTI,'YYYY-MM-DD') data
        FROM """ + S + """.ORD o JOIN """ + S + """.TIT t ON t.CODABI=o.CODABI
        WHERE o.CODCLI=:c AND TRUNC(o.DATSTI) BETWEEN TO_DATE(:dal,'YYYY-MM-DD') AND TO_DATE(:al,'YYYY-MM-DD')
        AND o.TIPOPE NOT LIKE 'FX%'
        ORDER BY o.DATSTI, t.DESTITB"""), {"c": codcli, "al": al, "dal": dal})
    return {
        "nav_al": float(sa["CONSFIN"]), "nav_dal": float(sd["CONSFIN"]),
        "tcli_al": float(sa["TCLI"]), "tcli_dal": float(sd["TCLI"]),
        "tbmk_al": float(sa["TBMK"]), "tbmk_dal": float(sd["TBMK"]), "bench": sa.get("CODIND") or "",
        "tcli_base": (float(bb["TCLI"]) if bb else None), "tbmk_base": (float(bb["TBMK"]) if bb else None),
        "nav_base": (float(bb["CONSFIN"]) if bb else None), "der_eq": der_eq, "deriv_pos": deriv_pos,
        "comp": [{"macro": r["MACRO"], "val_al": float(r["VAL_AL"] or 0), "val_dal": float(r["VAL_DAL"] or 0)} for r in comp],
        "trades": [{"nome": r["NOME"], "verso": r["VERSO"], "qty": float(r["QUANTI"] or 0),
                    "prezzo": float(r["PREZUNI"] or 0), "data": r["DATA"],
                    "ctv": float(r["QUANTI"] or 0) * float(r["PREZUNI"] or 0) * (0.01 if (r["G"] or "").startswith("A") else 1)} for r in trades],
    }

# ---------------- anagrafica contratto (per report cliente) ----------------
def contract_info(schema, codcli):
    """Ritorna {descli, tipoge, linea, gruppo} del contratto. DEMO: da contratti.json."""
    if mode() == "DEMO":
        try:
            for c in json.load(open(os.path.join(FIXTURES, "contratti.json"), encoding="utf-8")):
                if c.get("codcli") == codcli:
                    tg = c.get("tipoge") or ""
                    return {"descli": c.get("descli") or codcli, "tipoge": tg,
                            "linea": c.get("linea") or LINEA.get(tg, tg), "gruppo": _gruppo(c.get("schema"), tg)}
        except Exception:
            pass
        return {"descli": codcli, "tipoge": "", "linea": "", "gruppo": _gruppo(schema, "")}
    rows = _q("SELECT DESCLI, TIPOGE FROM " + schema + ".CNT WHERE CODCLI=:c ORDER BY DATCHIU NULLS FIRST FETCH FIRST 1 ROWS ONLY", {"c": codcli})
    if not rows:
        return {"descli": codcli, "tipoge": "", "linea": "", "gruppo": _gruppo(schema, "")}
    tg = rows[0].get("TIPOGE") or ""
    return {"descli": rows[0].get("DESCLI") or codcli, "tipoge": tg,
            "linea": LINEA.get(tg, tg), "gruppo": _gruppo(schema, tg)}

"""Checklist pre-esposizione (piano 3): auth, upload, riflessi HTML."""
import io

import openpyxl
import pytest

import data_layer as DL
from app import app as flask_app


@pytest.fixture()
def client():
    flask_app.config["TESTING"] = True
    return flask_app.test_client()


def test_con_auth_attiva_serve_l_header(client, monkeypatch):
    monkeypatch.setenv("COMITATO_AUTH", "1")
    assert client.get("/").status_code == 401
    r = client.get("/", headers={"X-Auth-Request-User": "mrossi@camperiosim.com"})
    assert r.status_code == 200


def test_static_resta_libero_con_auth_attiva(client, monkeypatch):
    monkeypatch.setenv("COMITATO_AUTH", "1")
    assert client.get("/static/style.css").status_code == 200


def test_upload_rifiuta_estensioni_non_ammesse(client, monkeypatch, tmp_path):
    monkeypatch.delenv("COMITATO_AUTH", raising=False)
    monkeypatch.setattr(DL, "REPO", str(tmp_path))
    r = client.post("/upload", data={"file": (io.BytesIO(b"testo"), "note.txt")},
                    content_type="multipart/form-data")
    d = r.get_json()
    assert d["ok"] is False and "Nessun file valido" in d["msg"]
    assert list(tmp_path.iterdir()) == []


def test_upload_riconosce_excel_fondo_dal_contenuto(client, monkeypatch, tmp_path):
    monkeypatch.delenv("COMITATO_AUTH", raising=False)
    monkeypatch.setattr(DL, "REPO", str(tmp_path))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "SUPERDISCOVERY UCITS"
    ws["B2"] = "14/08/2026"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = client.post("/upload", data={"file": (buf, "caricamento qualsiasi.xlsx")},
                    content_type="multipart/form-data")
    d = r.get_json()
    assert d["ok"] is True
    assert d["saved"] == ["SUPERDISCOVERY_2026-08-14.xlsx"]
    assert (tmp_path / "SUPERDISCOVERY_2026-08-14.xlsx").exists()


def test_parametro_al_ostile_viene_riflesso_escapato(client, monkeypatch):
    monkeypatch.delenv("COMITATO_AUTH", raising=False)
    r = client.get("/api/preview?schema=ANTASIMGEST&codcli=DEMO01&tipo=matrice"
                   "&al=<img src=x onerror=alert(1)>")
    html = r.get_json()["html"]
    assert "<img" not in html
    assert "&lt;img" in html


def test_header_legacy_fidato_dal_solo_nginx(client, monkeypatch):
    # L'app si fida anche di X-Remote-User: e' nginx a doverlo azzerare
    # (nginx.conf). Se questo test ti sorprende, aggiorna anche nginx.conf.
    monkeypatch.setenv("COMITATO_AUTH", "1")
    r = client.get("/", headers={"X-Remote-User": "chiunque"})
    assert r.status_code == 200
